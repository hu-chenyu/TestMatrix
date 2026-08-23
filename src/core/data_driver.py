"""
YAML/Excel数据驱动引擎模块

功能:
    - 统一入口 load_cases: 按文件后缀自动识别YAML/Excel，调用方无需关心底层格式
    - YAML解析: 支持顶层列表（parametrize直用格式）与顶层字典（唯一列表键）两种组织形式，
      嵌套结构字段原样保留
    - Excel解析: 基于openpyxl，支持指定sheet，首行作为字段名，自动跳过空行
    - 标准字段校验: case_id/name/module/priority/tags五字段强校验，
      错误提示携带行号（或用例序号）与字段名的中文定位信息
    - 用例筛选: filter_cases支持module（模块）/priority（优先级P0-P3）/tags（标签）三维过滤
    - 全链路日志: 加载、解析、校验、筛选每个环节均输出对应级别日志

使用示例:
    from src.core.data_driver import DataDriver

    # 统一入口加载（后缀自动识别）
    cases = DataDriver.load_cases("testdata/yaml/api_user_query_matrix.yaml")
    excel_cases = DataDriver.load_cases(
        "testdata/excel/api_user_query_matrix.xlsx", sheet_name="query_cases"
    )

    # 三维筛选（module精确匹配 / priority精确匹配 / tags任一命中即保留）
    smoke_cases = DataDriver.filter_cases(cases, tags=["smoke"])
    p0_cases = DataDriver.filter_cases(cases, priority="P0")

    # 直接驱动pytest参数化
    @pytest.mark.parametrize("case", cases, ids=[c["case_id"] for c in cases])
    def test_demo(case): ...
"""

from pathlib import Path
from typing import Any, Optional, Union

import yaml
from openpyxl import load_workbook

from src.common.logger import LogManager

logger = LogManager.get_logger()

# 项目根目录: 本模块位于 src/core/ 下，向上两级即为项目根
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent

# 支持的文件后缀（小写匹配）
YAML_SUFFIXES = (".yaml", ".yml")
EXCEL_SUFFIXES = (".xlsx",)

# 标准用例必填字段（缺失或为空即校验失败）
REQUIRED_FIELDS = ("case_id", "name", "module")
# 合法优先级取值
VALID_PRIORITIES = ("P0", "P1", "P2", "P3")


class DataDriverError(Exception):
    """
    数据驱动统一异常类

    封装数据文件加载、解析、校验、筛选过程中的异常，
    携带数据文件路径上下文，便于数据问题快速定位。
    """

    def __init__(self, message: str, file_path: Optional[Union[str, Path]] = None):
        """
        初始化异常

        参数:
            message (str): 异常描述信息
            file_path (str | Path | None): 出错的数据文件路径，用于日志排查

        返回:
            无
        """
        self.file_path = str(file_path) if file_path else None
        super().__init__(f"[数据文件 {self.file_path}] {message}" if self.file_path else message)


class DataDriver:
    """
    数据驱动引擎

    屏蔽YAML/Excel底层格式差异，对外提供统一的数据加载、
    校验与筛选能力，输出可直接用于pytest.mark.parametrize的列表结构。

    全部方法为类方法，无需实例化。
    """

    # ------------------------------------------------------------------
    # 统一加载入口
    # ------------------------------------------------------------------
    @classmethod
    def load_cases(
        cls,
        file_path: Union[str, Path],
        sheet_name: Optional[str] = None,
    ) -> list:
        """
        统一数据加载入口（按文件后缀自动分发到对应解析器）

        参数:
            file_path (str | Path): 数据文件路径；相对路径先按当前工作目录解析，
                                    不存在时再按项目根目录兜底解析
            sheet_name (str | None): Excel的sheet名称；None时使用活动sheet（仅Excel生效）

        返回:
            list[dict]: 校验通过并规范化后的用例列表，可直接用于pytest参数化

        异常:
            DataDriverError: 文件不存在 / 后缀不支持 / 格式解析失败 / 字段校验失败时抛出，
                             信息附带文件路径与行号（或用例序号）定位
        """
        path = cls._resolve_path(file_path)
        suffix = path.suffix.lower()
        logger.info(f"数据驱动加载开始 | 文件: {path} | 格式: {suffix}")

        if suffix in YAML_SUFFIXES:
            cases = cls._load_yaml(path)
        elif suffix in EXCEL_SUFFIXES:
            cases = cls._load_excel(path, sheet_name)
        else:
            raise DataDriverError(
                f"不支持的文件格式: '{suffix}'，"
                f"当前仅支持 {'/'.join(YAML_SUFFIXES + EXCEL_SUFFIXES)}",
                file_path=path,
            )

        # 逐条校验并规范化（校验失败抛出带定位信息的异常）
        validated_cases = [
            cls._validate_case(case, location=index, file_path=path)
            for index, case in enumerate(cases, start=1)
        ]

        logger.info(
            f"数据驱动加载完成 | 文件: {path.name} | 用例数: {len(validated_cases)} | "
            f"全部通过标准字段校验"
        )
        return validated_cases

    # ------------------------------------------------------------------
    # 用例筛选
    # ------------------------------------------------------------------
    @classmethod
    def filter_cases(
        cls,
        cases: list,
        module: Optional[Union[str, list]] = None,
        priority: Optional[Union[str, list]] = None,
        tags: Optional[Union[str, list]] = None,
    ) -> list:
        """
        按多维度筛选用例（未指定的维度不参与过滤，维度间为AND关系）

        匹配规则:
            - module   精确匹配（单值或列表任一命中）
            - priority 精确匹配、忽略大小写（单值或列表任一命中）
            - tags     交集匹配（用例tags与筛选tags存在任一交集即保留）

        参数:
            cases (list[dict]): 待筛选的用例列表（须为load_cases的规范化输出）
            module (str | list | None): 模块名，如"用户管理"或["用户管理", "订单"]
            priority (str | list | None): 优先级，如"P0"或["P0", "P1"]
            tags (str | list | None): 标签，如"smoke"或["smoke", "regression"]

        返回:
            list[dict]: 命中筛选条件的用例列表（浅拷贝，不影响原列表）

        异常:
            无（维度值为空列表时视为不过滤该维度并记录警告）
        """
        if not cases:
            logger.warning("用例筛选输入为空列表，直接返回空结果")
            return []

        # 统一归一化为列表便于匹配（空列表视为未指定该维度）
        module_list = cls._normalize_filter_value(module, "module")
        priority_list = cls._normalize_filter_value(priority, "priority")
        tags_list = cls._normalize_filter_value(tags, "tags")
        # 优先级匹配统一大写（与校验规范化的存储格式对齐）
        priority_list = [str(item).upper() for item in priority_list]

        matched = []
        for case in cases:
            if module_list and case.get("module") not in module_list:
                continue
            if priority_list and case.get("priority") not in priority_list:
                continue
            if tags_list and not set(case.get("tags", [])) & set(tags_list):
                continue
            matched.append(case)

        logger.info(
            f"用例筛选完成 | 输入: {len(cases)}条 | 输出: {len(matched)}条 | "
            f"条件: module={module_list or '-'} & priority={priority_list or '-'} & "
            f"tags={tags_list or '-'}"
        )
        return matched

    # ------------------------------------------------------------------
    # YAML解析
    # ------------------------------------------------------------------
    @classmethod
    def _load_yaml(cls, path: Path) -> list:
        """
        解析YAML数据文件（内部方法）

        支持两种顶层组织形式:
            1. 顶层列表: 直接作为用例列表（pytest参数化直用格式）
               - case_id: TM-0001
                 name: 示例
            2. 顶层字典: 唯一顶层键且值为列表时取该列表（分文件归类格式）
               query_cases:
                 - case_id: TM-0001

        参数:
            path (Path): YAML文件路径

        返回:
            list: 原始用例列表（未校验）

        异常:
            DataDriverError: 文件读取失败 / YAML语法错误 / 顶层结构非法 /
                             顶层字典存在多个列表键无法确定取值时抛出
        """
        try:
            with open(path, encoding="utf-8") as file_handle:
                data = yaml.safe_load(file_handle)
        except yaml.YAMLError as exc:
            logger.error(f"YAML语法解析失败 | {path} | {exc}")
            raise DataDriverError(f"YAML语法解析失败: {exc}", file_path=path) from exc
        except OSError as exc:
            logger.error(f"YAML文件读取失败 | {path} | {exc}")
            raise DataDriverError(f"文件读取失败: {exc}", file_path=path) from exc

        # 形式一: 顶层列表直接使用
        if isinstance(data, list):
            logger.debug(f"YAML解析完成[顶层列表格式] | {path.name} | 条数: {len(data)}")
            return data

        # 形式二: 顶层字典取唯一列表键
        if isinstance(data, dict):
            list_keys = [key for key, value in data.items() if isinstance(value, list)]
            if len(list_keys) == 1:
                logger.debug(
                    f"YAML解析完成[字典格式, 列表键: {list_keys[0]}] | "
                    f"{path.name} | 条数: {len(data[list_keys[0]])}"
                )
                return data[list_keys[0]]
            if not list_keys:
                raise DataDriverError(
                    "YAML顶层字典中未找到值为列表的键，无法提取用例数据", file_path=path
                )
            raise DataDriverError(
                f"YAML顶层字典存在多个列表键 {list_keys}，无法确定用例数据归属，"
                f"请拆分为多个数据文件或改为顶层列表格式",
                file_path=path,
            )

        raise DataDriverError(
            f"YAML顶层结构非法: {type(data).__name__}，仅支持列表或字典", file_path=path
        )

    # ------------------------------------------------------------------
    # Excel解析
    # ------------------------------------------------------------------
    @classmethod
    def _load_excel(cls, path: Path, sheet_name: Optional[str]) -> list:
        """
        解析Excel数据文件（内部方法）

        读取规则:
            - 首行为字段名（表头），其余行为用例数据
            - sheet_name为None时读取活动sheet，否则按名称读取指定sheet
            - 整行单元格均为空的行自动跳过
            - 空单元格对应的字段不写入用例字典（由字段校验兜底提示缺失）

        参数:
            path (Path): Excel文件路径（.xlsx）
            sheet_name (str | None): 指定sheet名称

        返回:
            list[dict]: 原始用例列表（未校验，行号信息由外层校验环节补齐）

        异常:
            DataDriverError: 文件打开失败 / 指定sheet不存在 / 表头为空时抛出
        """
        try:
            workbook = load_workbook(path, data_only=True, read_only=True)
        except Exception as exc:  # noqa: BLE001 openpyxl抛出类型多样，统一包装
            logger.error(f"Excel文件打开失败 | {path} | {exc}")
            raise DataDriverError(f"Excel文件打开失败: {exc}", file_path=path) from exc

        try:
            if sheet_name is None:
                worksheet = workbook.active
                actual_sheet = worksheet.title
            elif sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                actual_sheet = sheet_name
            else:
                raise DataDriverError(
                    f"指定sheet不存在: '{sheet_name}'，"
                    f"当前可用sheet: {workbook.sheetnames}",
                    file_path=path,
                )

            rows = list(worksheet.iter_rows(values_only=True))
            if not rows:
                raise DataDriverError("Excel内容为空（无表头行）", file_path=path)

            # 首行作为表头（strip清洗，空表头列及其数据列整体忽略）
            headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
            valid_columns = [(idx, header) for idx, header in enumerate(headers) if header]
            if not valid_columns:
                raise DataDriverError("Excel首行表头全部为空，无法识别字段名", file_path=path)

            cases = []
            skipped_rows = 0
            # 数据行从Excel第2行开始（row_index与Excel实际行号对齐，用于错误定位）
            for row_index, row in enumerate(rows[1:], start=2):
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    skipped_rows += 1
                    continue
                case = {}
                for col_index, header in valid_columns:
                    cell_value = row[col_index] if col_index < len(row) else None
                    if cell_value is not None:
                        case[header] = cell_value
                case["_row_number"] = row_index  # 行号暂存，校验环节用于错误定位
                cases.append(case)

            logger.info(
                f"Excel解析完成 | {path.name} | sheet: {actual_sheet} | "
                f"表头字段: {[header for _, header in valid_columns]} | "
                f"数据行: {len(cases)} | 跳过空行: {skipped_rows}"
            )
            return cases
        finally:
            workbook.close()

    # ------------------------------------------------------------------
    # 字段校验与规范化
    # ------------------------------------------------------------------
    @classmethod
    def _validate_case(cls, case: Any, location: int, file_path: Path) -> dict:
        """
        校验并规范化单条用例（内部方法）

        校验规则:
            - 用例必须为字典结构
            - case_id/name/module: 必填且为非空字符串
            - priority: 必填且必须为P0-P3（写入时统一大写规范化）
            - tags: 可选，列表或逗号分隔字符串（统一规范化为列表并strip）

        参数:
            case (Any): 单条原始用例数据
            location (int): 定位标识（YAML为用例序号，Excel为行号）
            file_path (Path): 数据文件路径（异常信息携带）

        返回:
            dict: 校验通过并规范化后的用例字典（priority大写、tags为列表）

        异常:
            DataDriverError: 任一校验规则不满足时抛出，
                             信息包含定位（行号/序号）与字段名
        """
        # Excel暂存行号优先作为定位信息（与Excel实际行号对齐）
        row_number = case.get("_row_number") if isinstance(case, dict) else None
        where = f"第{row_number}行" if row_number else f"第{location}条用例"

        if not isinstance(case, dict):
            raise DataDriverError(
                f"{where}结构非法: 期望字典，实际 {type(case).__name__}", file_path=file_path
            )

        validated = dict(case)
        validated.pop("_row_number", None)

        # 必填字段校验: 非空字符串
        for field in REQUIRED_FIELDS:
            value = validated.get(field)
            if not isinstance(value, str) or not value.strip():
                raise DataDriverError(
                    f"{where}必填字段'{field}'缺失或为空（要求非空字符串），"
                    f"当前值: {value!r}",
                    file_path=file_path,
                )

        # 优先级校验: 必须为P0-P3（大小写容错，统一大写回写）
        priority = validated.get("priority")
        if not isinstance(priority, str) or priority.strip().upper() not in VALID_PRIORITIES:
            raise DataDriverError(
                f"{where}字段'priority'非法: {priority!r}，合法取值: {list(VALID_PRIORITIES)}",
                file_path=file_path,
            )
        validated["priority"] = priority.strip().upper()

        # 标签校验: 列表或逗号分隔字符串，统一规范化为列表
        tags = validated.get("tags", [])
        if isinstance(tags, str):
            tags = [item.strip() for item in tags.split(",") if item.strip()]
        elif isinstance(tags, list):
            tags = [str(item).strip() for item in tags if str(item).strip()]
        else:
            raise DataDriverError(
                f"{where}字段'tags'非法: {tags!r}，要求列表或逗号分隔字符串",
                file_path=file_path,
            )
        validated["tags"] = tags

        logger.debug(
            f"用例校验通过 | {where} | case_id: {validated['case_id']} | "
            f"module: {validated['module']} | priority: {validated['priority']} | "
            f"tags: {validated['tags']}"
        )
        return validated

    # ------------------------------------------------------------------
    # 内部工具方法
    # ------------------------------------------------------------------
    @staticmethod
    def _resolve_path(file_path: Union[str, Path]) -> Path:
        """
        解析数据文件路径（内部方法）

        解析优先级: 绝对路径直接使用 > 相对路径按当前工作目录 >
                    相对路径按项目根目录兜底（保证任意cwd启动pytest均可定位）

        参数:
            file_path (str | Path): 原始文件路径

        返回:
            Path: 解析后的文件绝对路径

        异常:
            DataDriverError: 路径为空或所有候选位置均不存在时抛出
        """
        if not file_path or not str(file_path).strip():
            raise DataDriverError("数据文件路径不能为空")

        path = Path(file_path)
        if path.is_absolute():
            candidates = [path]
        else:
            candidates = [Path.cwd() / path, PROJECT_ROOT / path]

        for candidate in candidates:
            if candidate.exists():
                return candidate.resolve()

        raise DataDriverError(
            f"数据文件不存在: {file_path}，已尝试位置: "
            f"{[str(item) for item in candidates]}",
        )

    @staticmethod
    def _normalize_filter_value(value: Union[str, list, None], dim_name: str) -> list:
        """
        筛选维度值归一化为列表（内部方法）

        参数:
            value (str | list | None): 筛选值（单值/列表/None）
            dim_name (str): 维度名称（用于空列表警告日志）

        返回:
            list: 归一化后的列表；None或空列表返回空列表（表示不过滤该维度）

        异常:
            无
        """
        if value is None:
            return []
        if isinstance(value, str):
            return [value.strip()] if value.strip() else []
        if isinstance(value, list):
            if not value:
                logger.warning(f"筛选维度'{dim_name}'传入了空列表，视为不过滤该维度")
            return [str(item).strip() for item in value if str(item).strip()]
        logger.warning(f"筛选维度'{dim_name}'类型非法: {type(value).__name__}，视为不过滤该维度")
        return []
